from xml.dom.minidom import TypeInfo
from openpyxl import Workbook,load_workbook
import os

print(os.path.abspath(os.path.dirname(os.path.dirname(__file__))))


main_excel = os.path.join(os.path.abspath(os.path.dirname(os.path.dirname(__file__))),'source_data','main.xlsx')
detail_excel = os.path.join(os.path.abspath(os.path.dirname(os.path.dirname(__file__))),'source_data','detail.xlsx')


wb_main = load_workbook(main_excel)
ws_main = wb_main["股东名册"]


wb_detail = load_workbook(detail_excel)
ws_detail = wb_detail["股东明细"]


data_main = tuple(ws_main)
data_detail = tuple(ws_detail)


# 将目标表ws_detail转为字典

rows_list =[]
for row in ws_detail.rows:
    row_list = []
    for cell in row:
        row_list.append(cell.value)
    rows_list.append(row_list)
    
result = []
for i in range(len(rows_list) - 1):
    row_dict = {}
    for j in range(len(rows_list[0])):
        row_dict[rows_list[0][j]] = rows_list[i + 1][j]
    result.append(row_dict)

def read_excel():

    for j in range(1,len(data_main)):
        # print(data_main[j][0].value)
        code = data_main[j][0].value
        line = read_excel_detail(code)
        
        from excelcreate import create_workbook
        
        savepath = os.path.join(os.path.dirname(os.path.dirname(__file__)),'output')
        if not os.path.exists(savepath):
            os.mkdir(savepath)

        create_workbook(code,line,savepath)
        
        

def read_excel_detail(code):
            
    number = 0
    output_list = []
    for item in result:
        output_deatil = []
        # print(item['股东名册'])
        if item['股东名册']==code:
            number += 1
            if item['类型']=='个人':
                output_deatil.append("")    #提示
                output_deatil.append(number)    #序号
                output_deatil.append("个人")    #机构标识
                output_deatil.append(item['姓名'])    #股东简称
                output_deatil.append(item['姓名'])    #股东全称
                output_deatil.append(item['性别'])    #性别
                output_deatil.append(item['证件类型1'])    #证件类别
                output_deatil.append(item['身份证号'])    #证件号码
                output_deatil.append("")    #证件地址
                output_deatil.append("")    #法定代表人姓名
                output_deatil.append("")    #法定代表人证件类别
                output_deatil.append("")    #经办人姓名
                output_deatil.append("")    #法定代表人证件编号
                output_deatil.append("")    #经办人证件类别
                output_deatil.append("")    #经办人证件编号
                output_deatil.append("")    #联系地址
                output_deatil.append("")    #联系方式
                output_deatil.append("流通")    #股权性质
                output_deatil.append(item['持有数量(万元)']*10000)    #股权数量（股）
                output_deatil.append("")    #实缴数量
                # output_deatil.append("")    #股东性质
                
                if item['证件类型1']=='身份证':
                    output_deatil.append("境内自然人")
                else:
                    output_deatil.append("境外自然人")

                output_deatil.append("已确权")    #确权标识
                output_deatil.append("")    #控制股份数
                output_deatil.append("")    #开户银行
                output_deatil.append("")    #银行账号
                output_deatil.append("")    #实际入股金额
                output_deatil.append("")    #投票权数
                output_deatil.append("")    #持有其他股权所在银行名称
                output_deatil.append("")    #持有家数
                output_deatil.append("")    #在本企业派驻董监高人员情况
                output_deatil.append("")    #股权证书状态
                output_deatil.append("")    #外部账号
                output_deatil.append("")    #是否为内部职工
                output_deatil.append("")    #出资日期
                output_deatil.append("")    #是否为发起人股东
                output_deatil.append("")    #是否为控股股东
                output_deatil.append("")    #是否为主要股东
                output_deatil.append("")    #是否为董监高成员
                output_deatil.append("")    #备注

                output_list.append(output_deatil)
                
            elif item['类型']=='机构':
                output_deatil.append("")    #提示
                output_deatil.append(number)    #序号
                output_deatil.append("机构")    #机构标识
                output_deatil.append(item['企业简称'])    #股东简称
                output_deatil.append(item['企业全称'])    #股东全称
                output_deatil.append("")    #性别
                output_deatil.append(item['证件类型2'])    #证件类别
                output_deatil.append(item['统一社会信用代码'])    #证件号码
                output_deatil.append("")    #证件地址
                output_deatil.append("")    #法定代表人姓名
                output_deatil.append("")    #法定代表人证件类别
                output_deatil.append("")    #经办人姓名
                output_deatil.append("")    #法定代表人证件编号
                output_deatil.append("")    #经办人证件类别
                output_deatil.append("")    #经办人证件编号
                output_deatil.append("")    #联系地址
                output_deatil.append("")    #联系方式
                output_deatil.append("流通")    #股权性质
                output_deatil.append(item['持有数量(万元)']*10000)    #股权数量（股）
                output_deatil.append("")    #实缴数量
                output_deatil.append("")    #股东性质
                output_deatil.append("已确权")    #确权标识
                output_deatil.append("")    #控制股份数
                output_deatil.append("")    #开户银行
                output_deatil.append("")    #银行账号
                output_deatil.append("")    #实际入股金额
                output_deatil.append("")    #投票权数
                output_deatil.append("")    #持有其他股权所在银行名称
                output_deatil.append("")    #持有家数
                output_deatil.append("")    #在本企业派驻董监高人员情况
                output_deatil.append("")    #股权证书状态
                output_deatil.append("")    #外部账号
                output_deatil.append("")    #是否为内部职工
                output_deatil.append("")    #出资日期
                output_deatil.append("")    #是否为发起人股东
                output_deatil.append("")    #是否为控股股东
                output_deatil.append("")    #是否为主要股东
                output_deatil.append("")    #是否为董监高成员
                output_deatil.append("")    #备注

                
                output_list.append(output_deatil)

    print('代码:',code,'\n股东人数:',number,'\n输出值:',output_list,'\n')
    return output_list

read_excel()
    